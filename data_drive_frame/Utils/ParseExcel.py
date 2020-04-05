#encoding:utf-8
from openpyxl import load_workbook

class ParseExcel(object):
    def __init__(self,excel_file_path):
        try:
            self.excel_file_path = excel_file_path
            self.wb = load_workbook(excel_file_path)
            self.sheet = self.wb[self.wb.sheetnames[0]]
        except Exception as e:
            raise e

    #获取测试数据
    def set_sheet_by_name(self,sheet_name):
        try:
            self.sheet = self.wb[sheet_name]
        except Exception as e:
            raise e

    #获取所有的行
    def get_all_rows(self):
        try:
            rows = list(self.sheet.rows)
            return rows
        except Exception as e:
            raise e


if __name__ == '__main__':
    pe = ParseExcel(r"C:\Users\pei.lu.KINGSTAR\eclipse-workspace\new_2020\data_drive_frame\TestData\测试数据.xlsx")
    print(pe.sheet.title)
    pe.set_sheet_by_name("发送邮件数据A")
    rows = pe.get_all_rows()
    for row in rows:
        print(row)


