# -*- coding: utf-8 -*-
"""
@time: 2020/2/20 9:36
@author: pei.lu
"""
# python中与excel操作相关的模块：
#
# xlrd库：从excel中读取数据，支持xls、xlsx
# xlwt库：对excel进行修改操作，不支持对xlsx格式的修改
# xlutils库：在xlwt和xlrd中，对一个已存在的文件进行修改。
# openpyxl：主要针对xlsx格式的excel进行读取和编辑。

# 1.1 xlrd读excel
def excel_1():
    import xlrd
    file = "excel_dxtest.xlsx"
    book=xlrd.open_workbook(file)
    # sheet1=book.sheets()[1]
    sheetnum=book.nsheets # 有几个sheet页
    sheet1=book.sheet_by_name('test_2')
    nrows,ncols=sheet1.nrows,sheet1.ncols # sheet页的总行数、总列数
    row3_values=sheet1.row_values(2) # 返回第3行值
    col3_values=sheet1.col_values(2) # 返回第3列值
    cell_3_3=sheet1.cell_value(2,2) # 返回第3行第3列单元格的值
    print(book,sheetnum,sheet1,nrows,row3_values,cell_3_3,sep='\n')

# excel_1()

# 1.2 xlwt写excel
def excel_2():
    import xlwt
    workbook=xlwt.Workbook()
    worksheet=workbook.add_sheet('test')
    worksheet.write(0,0,'A1data')
    worksheet.write(1,2,'中文')
    worksheet.write(2,0,'wei=tsdf133')
    workbook.save('excelwrite.xls')

# 读文件后写新文件
def excel_3():
    import xlrd,xlwt
    file='excel_dxtest.xlsx'
    wb=xlrd.open_workbook(file)
    sh=wb.sheet_by_name('test_2')
    col5=sh.col_values(5)
    new_list=['data='+x for x in col5[1:]]
    # print(new_list)
    new_wb=xlwt.Workbook()
    new_sh=new_wb.add_sheet('gp2')
    for i,value in enumerate(new_list):
        new_sh.write(i,5,value)
    new_wb.save('new_data.xls')

# excel_3()

def excel_openpyxl_1():
    import openpyxl
    # 创建一个工作簿
    wb = openpyxl.Workbook()
    # 创建一个test_case的sheet表单
    sh=wb.create_sheet('test_case')
    sh1=wb.create_sheet('test_case1')
    wb.remove(sh1)  # 删除表单
    sh.cell(1,1,value='A1')
    sh.cell(2,1,'A2')
    sh.cell(3,1,'A3')
    sh.cell(1,4,value='A4')
    sh['B1']='hi,B1'
    sh['B2'] = 'B2cell'
    sh['B3'] = 'B3cell'
    sh['C1']='中国馆'
    # 保存为一个xlsx格式的文件
    wb.save(r'common_file\cases.xlsx')
# excel_openpyxl_1()

def excel_openpyxl_2():
    import openpyxl
    # 第一步：打开工作簿
    wb = openpyxl.load_workbook(r'cases.xlsx')
    # 第二步：选取表单
    sh = wb['test_case']
    # 第三步：读取数据
    ce = sh.cell(row=1, column=1)  # 读取第一行，第一列的数据
    print(ce.value)
    print(sh.max_row, sh.max_column)  # 最大行、最大列总数
    # 按行读取数据 list(sh.rows)
    print(list(sh.rows)[1:])  # 按行读取数据，去掉第一行的表头信息数据
    for cases in list(sh.rows)[1:]:
        case_id = cases[0].value
        case_excepted = cases[1].value
        case_data = cases[2].value
        print(case_id,case_excepted, case_data)
    # 关闭工作薄
    wb.close()

excel_openpyxl_2()