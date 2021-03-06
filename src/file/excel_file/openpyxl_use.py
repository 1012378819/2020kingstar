# -*- coding: utf-8 -*-
"""
@time: 2020/3/1 11:13
@author: pei.lu
"""
#  https://www.cnblogs.com/wanglle/p/11455758.html
import openpyxl
class Case:
    # __slots__ = [] # 特殊的类属性，用来限制类创建实例属性添加
    pass

class ReadExcel(object):
    def __init__(self,excel_name,sheet_name):
        """
        :param excel_name:文件名  -->str类型
        :param sheet_name:表单名  -->str类型
        """
        self.wb=openpyxl.load_workbook(excel_name)
        self.sh=self.wb[sheet_name]

    # 按行读取，存储在列表中
    def read_data_line(self):
        # 按行读取数据转为list
        rows_data=list(self.sh.rows)
        titles=[]
        for title in rows_data[0]:
            titles.append(title.value)
        cases=[]
        for case in rows_data[1:]:
            data=[]
            for cell in case:
                # data.append(cell.value)
                # 判断是否为字符串，如果是字符串类型则需要使用eval();如果不是字符串类型则不需要使用eval()
                if isinstance(cell.value,str):
                    data.append(eval(cell.value))
                else:
                    data.append(cell.value)
            case_data=dict(list(zip(titles,data)))
            cases.append(case_data)
        return cases

    # 按行读取，存储在对象中
    def read_data_obj(self):
        rows_data=list(self.sh.rows)
        titles=[]
        for title in rows_data[0]:
            titles.append(title.value)
        cases=[]
        for case in rows_data[1:]:
            case_obj=Case() # 创建一个Case类对象，用来保存用例数据
            data=[]
            for cell in case:
                if isinstance(cell.value,str):
                    data.append(eval(cell.value))
                else:
                    data.append(cell.value)
            case_data=list(zip(titles,data))
            for i in case_data:
                setattr(case_obj,i[0],i[1])
            cases.append(case_obj)
        return cases

    # 将测试用例封装到列表中，读取指定列的数据
    def read_data(self,list1):
        """
        :param list1:  list--->要读取列   list类型
        :return: 返回一个列表，每一个元素为一个用例（用例为dict类型）
        """
        max_r=self.sh.max_row
        cases=[] # 存放所有用例数据
        titles=[] # 存放表头
        for row in range(1,max_r+1):
            if row!=1:
                case_data=[]
                for col in list1:
                    info=self.sh.cell(row,col).value
                    case_data.append(info)
                case=dict(zip(titles,case_data))
                cases.append(case)
            else:
                for col in list1:
                    title=self.sh.cell(row,col).value
                    titles.append(title)
        return cases

    # 将测试用例封装到对象中，读取指定列的数据
    def read_data_obj_optimize(self,list2):
        max_r = self.sh.max_row
        cases = []  # 存放所有用例数据
        titles = []  # 存放表头
        for row in range(1, max_r + 1):
            if row != 1:
                case_data = []
                for col in list2:
                    info = self.sh.cell(row, col).value
                    case_data.append(info)
                case = list(zip(titles, case_data))
                case_obj=CaseOptimize(case)
                cases.append(case_obj)
            else:
                for col in list2:
                    title = self.sh.cell(row, col).value
                    titles.append(title)
                if None in titles:
                    raise ValueError("传入的表头的数据有显示为空")
        return cases

class CaseOptimize: # 优化
    def __init__(self,attrs):
        for i in attrs:
            setattr(self,i[0],i[1])

# excel_3()
def excel_openpyxl_1():
    import openpyxl
    # 创建一个工作簿
    wb = openpyxl.Workbook()
    # 创建一个test_case的sheet表单
    sh=wb.create_sheet('test_case')
    sh1=wb.create_sheet('test_case1')
    wb.remove(sh1)  # 删除表单
    sh.cell(1,1,value='A1')
    sh.cell(2,1,'A2')
    sh.cell(3,1,'A3')
    sh.cell(1,4,value='A4')
    sh['B1']='hi,B1'
    sh['B2'] = 'B2cell'
    sh['B3'] = 'B3cell'
    sh['C1']='中国馆'
    # 保存为一个xlsx格式的文件
    wb.save(r'common_file\cases.xlsx')
# excel_openpyxl_1()

def excel_openpyxl_2():
    import openpyxl
    # 第一步：打开工作簿
    wb = openpyxl.load_workbook(r'cases.xlsx')
    # 第二步：选取表单
    sh = wb['test_case']
    # 第三步：读取数据
    ce = sh.cell(row=1, column=1)  # 读取第一行，第一列的数据
    print(ce.value)
    print(sh.max_row, sh.max_column)  # 最大行、最大列总数
    # 按行读取数据 list(sh.rows)
    print(list(sh.rows)[1:])  # 按行读取数据，去掉第一行的表头信息数据
    for cases in list(sh.rows)[1:]:
        case_id = cases[0].value
        case_excepted = cases[1].value
        case_data = cases[2].value
        print(case_id,case_excepted, case_data)
    # 关闭工作薄
    wb.close()

excel_openpyxl_2()

if __name__ == '__main__':
    f = 'common_file\\apicases.xlsx'
    r=ReadExcel(f,'sheet1')
    data=r.read_data_line()
    data1=r.read_data_obj()
    print(data)
    for i in data1:
        print(i.caseid,i.data,i.expected)
    res=r.read_data([1,2,3])
    for o in res:
        print(o['caseid'],o['data'],o['expected'])
    res1=r.read_data_obj_optimize([1,2,3])
    for i in res1:
        print(i.caseid,i.data,i.expected)
